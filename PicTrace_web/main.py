import os
import time
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from logger import setup_logging
from config import IMAGE_FOLDER, RESULTS_FILE, MAX_TOTAL_RECORDS
from browser import setup_browser
from image_matcher import ImageMatcher
from search import search_yandex_image

def process_images():
    setup_logging()
    if not os.path.exists(IMAGE_FOLDER):
        os.makedirs(IMAGE_FOLDER)
        logging.info(f"Created folder: {IMAGE_FOLDER}")
        return
    matcher = ImageMatcher()
    driver = setup_browser()
    all_data = []
    record_counter = [0]
    try:
        for image_name in os.listdir(IMAGE_FOLDER):
            if image_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                image_path = os.path.join(IMAGE_FOLDER, image_name)
                logging.info(f"Processing: {image_name}")
                source_embedding = matcher.get_embedding(image_path)
                if source_embedding is None:
                    logging.warning(f"Failed to extract embedding for: {image_name}. Skipping.")
                    continue
                logging.info(f"Extracted embedding for: {image_name}")
                logging.info("Starting search on Yandex.")
                search_yandex_image(driver, image_path, matcher, source_embedding, image_name, all_data, record_counter)
                if record_counter[0] >= MAX_TOTAL_RECORDS:
                    logging.info(f"Reached maximum number of records: {MAX_TOTAL_RECORDS}")
                    break
                time.sleep(2)
        if all_data:
            df = pd.DataFrame(all_data)
            excel_file = RESULTS_FILE
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Results')
                workbook = writer.book
                worksheet = writer.sheets['Results']
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    column_letter = cell.column_letter
                    max_length = max(df.iloc[:, col - 1].astype(str).map(len).max(), len(cell.value))
                    worksheet.column_dimensions[column_letter].width = max_length + 2
                if 'Similarity' in df.columns:
                    similarity_col = df.columns.get_loc('Similarity') + 1
                    for row in range(2, len(df) + 2):
                        similarity_value = df.iloc[row - 2]['Similarity']
                        try:
                            similarity_percent = float(similarity_value)
                        except:
                            similarity_percent = 0.0
                        if similarity_percent >= 0.85:
                            color_intensity = min(int((similarity_percent - 0.85) / 0.15 * 255), 255)
                            fill_color = f"{255:02X}{255 - color_intensity:02X}{255 - color_intensity:02X}"
                            cell = worksheet.cell(row=row, column=similarity_col)
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell = worksheet.cell(row=row, column=similarity_col)
                        cell.border = thin_border
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                # Adding separators between different images
                if 'Image Name' in df.columns:
                    last_image = None
                    for row in range(2, len(df) + 2):
                        current_image = df.iloc[row - 2]['Image Name']
                        if last_image and current_image != last_image:
                            for col in range(1, len(df.columns) + 1):
                                divider_cell = worksheet.cell(row=row, column=col)
                                divider_cell.fill = PatternFill(start_color="000000", end_color="000000",
                                                                fill_type="solid")
                        last_image = current_image
                # Adding stripes every 10 rows
                for row in range(2, len(df) + 2):
                    if row % 10 == 0:
                        for col in range(1, len(df.columns) + 1):
                            divider_cell = worksheet.cell(row=row, column=col)
                            divider_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            logging.info(f"Data successfully written to {excel_file}")
        else:
            logging.info("No data to write.")
    except Exception as e:
        logging.error(f"Fatal error: {str(e)}")
    finally:
        driver.quit()
        logging.info("Browser closed")


if __name__ == "__main__":
    process_images()
